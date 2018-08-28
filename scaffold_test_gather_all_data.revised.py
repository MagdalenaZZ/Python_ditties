#!/usr/bin/env python3.3

import sys
import argparse
import os
import itertools
import fastn
import utils
sys.path.insert(1, '/nfs/users/nfs_m/mh12/lib/python3.2/openpyxl-1.6.2/')
import openpyxl
import fractions
import functools

parser = argparse.ArgumentParser(
    description = 'Used to gather data for scaffolder testing. Input directories/files are hard-coded!',
    usage = '%(prog)s [options] <extra cpu file> <output_prefix>')
parser.add_argument('--use_sspace_iter', action='store_true', help='Use this to include sspace_iter results')
parser.add_argument('extra_cpu_file', help='File of: data tool extra_cpu extra_mem')
parser.add_argument('outprefix', help='Prefix of output files')
options = parser.parse_args()



datasets = [
    'Staph.perfect.3kb.short',
    'Staph.perfect.3kb.long',
    'Staph.perfect.10kb.short',
    'Staph.perfect.10kb.long',
    'Staph.velvet',
    'Rhodobacter',
    '3D7.short',
    '3D7.long',
    '3D7.combi',
    'human.short',
    'human.long',
    'human.combi'
]

root_dir = '/lustre/scratch108/parasites/mh12/Scaffold_test'

dataset2dir = {
    'Staph.perfect.3kb.short': 'GAGE-staph-perfect/Scaffold.3kb-contigs.500bp-reads',
    'Staph.perfect.3kb.long': 'GAGE-staph-perfect/Scaffold.3kb-contigs.3kb-reads',
    'Staph.perfect.10kb.short': 'GAGE-staph-perfect/Scaffold.10kb-contigs.3kb-reads',
    'Staph.perfect.10kb.long': 'GAGE-staph-perfect/Scaffold.10kb-contigs.500bp-reads',
    'Staph.velvet': 'GAGE-staph-velvet/Scaffold',
    'Rhodobacter': 'GAGE-rhodobacter/Scaffold',
    '3D7.short': '3D7/velvet-k55/Scaffold.5853_6',
    '3D7.long': '3D7/velvet-k55/Scaffold.7101_2-4',
    '3D7.combi': '3D7/velvet-k55/Scaffold.5853_6_and_7101_2-4',
    'human.short': 'GAGE-human-chr14/Scaffold.shortjump',
    'human.long': 'GAGE-human-chr14/Scaffold.longjump.trimmed',
    'human.combi': 'GAGE-human-chr14/Scaffold.short_then_long.trimmed'
}

scaffolders = [
    'ABySS',
    'Bambus2.bowtie2',
    'Bambus2.bwa',
    'MIP.bowtie-v0',
    'MIP.bowtie-v3',
    'MIP.bowtie2',
    'MIP.bwa',
    'OPERA.bowtie',
    'OPERA.bwa',
    'SCARPA.bowtie-v0',
    'SCARPA.bowtie-v3',
    'SCARPA.bowtie2',
    'SCARPA.bwa',
    'SGA.bowtie2',
    'SGA.bwa',
    #'SOAP',
    'SOAP2',
    'SOPRA.bowtie-v0',
    'SOPRA.bowtie-v3',
    'SOPRA.bowtie2',
    'SOPRA.bwa',
    'SSPACE.bowtie-v0',
    'SSPACE.bowtie-v3'
]


r_colours = {
    'ABySS': 'cyan2',
    'Bambus2.bowtie2': 'aquamarine4',
    'Bambus2.bwa': 'aquamarine4',
    'MIP.bowtie-v0': 'gray',
    'MIP.bowtie-v3': 'gray',
    'MIP.bowtie2': 'gray',
    'MIP.bwa': 'gray',
    'OPERA.bowtie': 'red',
    'OPERA.bwa': 'red',
    'SCARPA.bowtie-v0': 'blue',
    'SCARPA.bowtie-v3': 'blue',
    'SCARPA.bowtie2': 'blue',
    'SCARPA.bwa': 'blue',
    'SGA.bowtie2': 'green',
    'SGA.bwa': 'green',
    'SOAP': 'purple',
    'SOAP2': 'purple',
    'SOPRA.bowtie-v0': 'brown',
    'SOPRA.bowtie-v3': 'brown',
    'SOPRA.bowtie2': 'brown',
    'SOPRA.bwa': 'brown',
    'SSPACE.bowtie-v0': 'orange',
    'SSPACE.bowtie-v3': 'orange'}

if options.use_sspace_iter:
    scaffolders.append('SSPACE.iter')
    r_colours['SSPACE.iter'] = 'orange'

class ScaffResults:
    possible_flags = [0,1,2,4,5,8,12,16]
    evaluation_score_keys = [
        'Correct joins',
        'Bad joins',
        'Lost tags',
        'Skipped tags',
        'Total CPU']

    headers = ['Dataset', 'Scaffolder', 'Correct joins'] \
              + [str(x) for x in possible_flags[1:-1]] + [
                'Bad joins',
                'Lost tags',
                'Skipped tags',
                'CPU',
                'Memory',
                'Extra CPU',
                'Extra Memory',
                'Total CPU',
                'Max memory',
                'Correct joins score',
                'Bad joins score',
                'Lost tags score',
                'Skipped tags score',
                'Total CPU score']

    def __init__(self, bsub_o, log_file, extra_cpu=0, extra_mem=0):
        # get flag counts etc from the log file
        self.flag_counts = {k:0 for k in ScaffResults.possible_flags}
        self.stats = {k: 0 for k in ScaffResults.evaluation_score_keys}

        if os.path.exists(log_file):
            f = utils.open_file_read(log_file)
            for line in f:
                a = line.split()

                if a[0].isdigit():
                    self.flag_counts[int(a[0])] = int(a[1])
                elif a[0] == 'lost':
                    self.stats['Lost tags'] = int(a[1])
                elif a[0] == 'skipped':
                    self.stats['Skipped tags'] = int(a[1])
            utils.close(f)
            self.stats['Bad joins'] = sum([self.flag_counts[x] for x in self.flag_counts.keys() if x not in [0,16]]) + self.stats['Lost tags']
        else:
            print('Warning: no log file', log_file, file=sys.stderr)

        # get cpu and mem from bsub file
        bsub_out = utils.syscall_get_stdout('bsub-out2stats.py -s ' + bsub_o)
        assert len(bsub_out) == 1
        (attempt_no, exit_code, wall_hrs, cpu_secs, cpu_hrs, mem, swap, filename) = bsub_out[0].split('\t')
        assert exit_code == '0'

        self.stats['Correct joins'] = self.flag_counts[0]
        self.cpu = int(round(float(cpu_secs), 0))
        self.mem = int(mem)
        self.extra_cpu = extra_cpu
        self.extra_mem = extra_mem
        self.stats['Total CPU'] = self.cpu + extra_cpu
        self.max_mem = max(self.mem, extra_mem)
        self.scores = {k:-1 for k in ScaffResults.evaluation_score_keys}
        self.worksheet_row = -1

    def correct_cpu(self, secs):
        self.cpu -= secs
        self.stats['Total CPU'] = self.cpu + self.extra_cpu

    def to_tsv(self):
        l = [self.flag_counts[x] for x in ScaffResults.possible_flags if x != 16]
        l += [self.stats['Bad joins'],
                self.stats['Lost tags'],
                self.stats['Skipped tags'],
                self.cpu,
                self.mem,
                self.extra_cpu,
                self.extra_mem,
                self.stats['Total CPU'],
                self.max_mem,
                round(self.scores['Correct joins'], 4),
                round(self.scores['Bad joins'], 4),
                round(self.scores['Lost tags'], 4),
                round(self.scores['Skipped tags'], 4),
                round(self.scores['Total CPU'], 4)]

        return '\t'.join([str(x) for x in l])


bad_runs = set()
#bad_runs.add(('3D7.combi', 'ABySS'))
bad_runs.add(('3D7.combi', 'Bambus2.bowtie2'))
bad_runs.add(('3D7.combi', 'Bambus2.bwa'))
#bad_runs.add(('human.long', 'ABySS'))
bad_runs.add(('human.long', 'SOPRA.bowtie2'))
#bad_runs.add(('human.combi', 'ABySS'))
bad_runs.add(('human.combi', 'MIP.bowtie2'))
bad_runs.add(('human.combi', 'MIP.bwa'))
bad_runs.add(('human.combi', 'SOPRA.bowtie2'))
bad_runs.add(('human.combi', 'SOPRA.bwa'))

# had to suspend a few jobs because lustre was taken down. So need
# to correct the CPU time
cpu_time_susp = {
    ('human.short', 'SOPRA.bowtie2'): 13608,
    ('human.combi', 'SOPRA.bowtie-v3'): 13608,
    ('human.combi', 'MIP.bwa'): 13990
}

data_row_ranges = {}
excel_outfile = options.outprefix + '.xlsx'
workbook = openpyxl.Workbook()
worksheet_all = workbook.worksheets[0]
worksheet_all.title = 'All results'
worksheet_scores = workbook.create_sheet()
worksheet_scores.title = 'Scores'

f_tsv = utils.open_file_write(options.outprefix + '.tsv')
print('\t'.join(ScaffResults.headers), file=f_tsv)

eval_keys_columns = {}

for i in range(len(ScaffResults.headers)):
    h = ScaffResults.headers[i]
    worksheet_all.cell(row=0, column=i).value = h
    if h.endswith(' score') and h[:-6] in ScaffResults.evaluation_score_keys:
        eval_keys_columns[h] = i


current_row = 1



# get the extra cpu and mem usage
extra_cpu = {}
extra_mem = {}

f = utils.open_file_read(options.extra_cpu_file)

for line in f:
    if line.startswith('#'):
        continue
    (dataset, scaffolder, cpu, mem) = line.split('\t')
    mem = int(mem)
    cpu = int(cpu)
    extra_cpu[(dataset, scaffolder)] = cpu
    extra_mem[(dataset, scaffolder)] = mem

utils.close(f)

# gather all the counts for each scaffolding run
results = {k:{} for k in datasets}

for dataset in datasets:
    #if 'human' in dataset:
    #    continue
    d = results[dataset]
    dataset_dir = os.path.join(root_dir, dataset2dir[dataset])

    for scaff in scaffolders:
        if (dataset, scaff) in bad_runs:
            continue
        print(dataset, dataset_dir, scaff, sep='\t')
        scaff_dir = os.path.join(dataset_dir, scaff)
        log_file = os.path.join(scaff_dir, 'check_scaffolds.log')
        bsub_outfile = os.path.join(scaff_dir, scaff.split('.')[0].lower() + '.o')
        if (dataset, scaff) in extra_cpu:
            d[scaff] = ScaffResults(bsub_outfile, log_file, extra_cpu=extra_cpu[(dataset, scaff)], extra_mem=extra_mem[(dataset, scaff)])
        else:
            d[scaff] = ScaffResults(bsub_outfile, log_file)

        if (dataset, scaff) in cpu_time_susp:
            d[scaff].correct_cpu(cpu_time_susp[(dataset, scaff)])

    # work out the scores for this dataset
    for score in ScaffResults.evaluation_score_keys:
        min_score = min([d[scaff].stats[score] for scaff in scaffolders if (dataset, scaff) not in bad_runs])
        max_score = max([d[scaff].stats[score] for scaff in scaffolders if (dataset, scaff) not in bad_runs])

        def get_score(lowest, highest, val, flip=False):
            if lowest == highest:
                return 1
            else:
                x = float(val - lowest) / float(highest - lowest)
                if flip:
                    return 1 - x
                else:
                    return x

        flip = False

        if score != 'Correct joins':
            flip = True

        for scaff in scaffolders:
            if (dataset, scaff) in bad_runs:
                continue
            d[scaff].scores[score] = get_score(min_score, max_score, d[scaff].stats[score], flip)

    data_range_start = current_row

    for scaff in scaffolders:
        if (dataset, scaff) in bad_runs:
            continue
        data_to_print = [dataset, scaff] +  d[scaff].to_tsv().split('\t')
        print('\t'.join(data_to_print), file=f_tsv)
        worksheet_all.append(data_to_print)
        d[scaff].worksheet_row = current_row
        current_row += 1

    data_row_ranges[dataset] = (data_range_start, current_row - 1)

utils.close(f_tsv)


# make the second excel sheet with score summary etc
worksheet_scores.cell(row=0, column=0).value = 'Weights'
current_row = 1
weights_rows = {}

for k in ScaffResults.evaluation_score_keys:
    worksheet_scores.append([k, 1])
    weights_rows[k] = current_row
    current_row += 1

worksheet_scores.append([''])
worksheet_scores.append([''])
worksheet_scores.append([''])

current_row += 4
worksheet_scores.append([''] + datasets + ['Total (real only)', 'Total (all)'])
current_row

for i in range(current_row, current_row + len(scaffolders)):
    worksheet_scores.cell(row=i, column=0).value = scaffolders[i - current_row]

start_row = current_row


to_print = {x:[x] for x in scaffolders}


current_column = 0

# make the summary for each dataset/tool
for col in range(len(datasets)):
    dataset = datasets[col]
    #if 'human' in dataset:
    #    for x in to_print.values():
    #        x.append(0)

    #    continue

    current_row = start_row
    current_column += 1
    column_to_print = []

    for scaff in scaffolders:
        if (dataset, scaff) in bad_runs:
            worksheet_scores.cell(row=current_row, column=current_column).value = "NA"
            current_row += 1
            continue

        # get the weighted score for this datset/test
        l = []

        for key in ScaffResults.evaluation_score_keys:
            weight_cell = worksheet_scores.cell(row=weights_rows[key], column=1).address
            data_row = results[dataset][scaff].worksheet_row
            l.append(weight_cell + " * 'All results'!" + worksheet_all.cell(row=data_row, column=eval_keys_columns[key + ' score']).address)

        formula = '=(' + ' + '.join(l) + ') / sum(b2:b6)'
        worksheet_scores.cell(row=current_row, column=current_column).value = formula
        current_row += 1

workbook.save(filename=excel_outfile)


# work out the scores for a range of weights
scores = {k:{} for k in datasets}
bins = 20

weight_ranges = {
        'Correct joins': [10,20,40,80,160],
        'Bad joins': [10,20,40,80,160],
        'Lost tags': [20,40,80,160],
        'Skipped tags': [10,20,40,80,160],
        'Total CPU': [1,2,3,4,5]
}

# want to add a star on each boxplot where these weightings would put the score
weights_for_starring = (80,160,160,40,1)
boxplot_stars = {d:{s:2 for s in scaffolders} for d in datasets}

#boxplot_scores = {x:{y:[] for y in scaffolders if (x,y) not in bad_runs} for x in datasets}
boxplot_scores = {x:{y:[] for y in scaffolders} for x in datasets}

for dataset in scores:
    for scaff in scaffolders:
        if (dataset, scaff) in bad_runs:
            continue
        scores[dataset][scaff] = [0] * (bins + 1)

used_weights = set()

# loop over all possible values of weights.
# Don't repeat duplicate weight combinations.
# e.g. doubling all the weights would be the same
for iter in itertools.product(*[weight_ranges[k] for k in ScaffResults.evaluation_score_keys]):
    weights = list(iter)

    d = {ScaffResults.evaluation_score_keys[i]: weights[i] for i in range(len(weights))}
    #only want:
    #   bad joins >= good joins
    #   bad joins >= 2 * skipped tags
    #   good joins <= lost tags
    if d['Bad joins'] < d['Correct joins'] or d['Bad joins'] < 2 * d['Skipped tags'] or d['Correct joins'] > d['Lost tags']:
        continue


    gcd = functools.reduce(fractions.gcd,weights)
    weights = [x / gcd for x in weights]

    assert len(weights) == len(ScaffResults.evaluation_score_keys)

    if (tuple(weights)) in used_weights:
        continue

    used_weights.add(tuple(weights))

    for dataset in datasets:
        #if 'human' in dataset:
        #    continue
        for scaff in scaffolders:
            if (dataset, scaff) in bad_runs:
                boxplot_scores[dataset][scaff].append(10)
                continue
            score = 0

            for eval_index in range(len(ScaffResults.evaluation_score_keys)):
                eval_key = ScaffResults.evaluation_score_keys[eval_index]
                weight = weights[eval_index]
                score += weight * results[dataset][scaff].scores[eval_key]

            score /= sum(weights)
            scores[dataset][scaff][int(score * bins)] += 1
            boxplot_scores[dataset][scaff].append(score)

            if tuple(weights) == weights_for_starring:
                boxplot_stars[dataset][scaff] = score


summary_datasets = {'all': datasets,
    'real_only': [x for x in datasets if 'perfect' not in x],
    'sim_only': [x for x in datasets if 'perfect' in x]
}







# Make a plot of score boxplots, for each dataset.
for dataset in datasets:
    #if 'human' in dataset:
    #    continue
    #r_colours_string = 'c(' + ','.join(['"' + r_colours[scaff] + '"' for scaff in scaffolders if (dataset, scaff) not in bad_runs]) + ')'
    r_colours_string = 'c(' + ','.join(['"' + r_colours[scaff] + '"' for scaff in scaffolders]) + ')'
    #r_scaffnames_string = 'c(' + ','.join(['"' + x + '"' for x in scaffolders if (dataset, x) not in bad_runs]) + ')'
    r_scaffnames_string = 'c(' + ','.join(['"' + x + '"' for x in scaffolders]) + ')'
    r_stars = 'c(' + ','.join([str(boxplot_stars[dataset][x]) for x in scaffolders]) + ')'
    r_script = options.outprefix + '.boxplot.' + dataset + '.R'
    f = utils.open_file_write(r_script)
    scaffnames = []

    for i in range(len(scaffolders)):
        #if (dataset, scaffolders[i]) in bad_runs:
        #    continue

        scaffname = 's' + str(i)
        scaffnames.append(scaffname)
        print(scaffname, '=c(', ','.join([str(x) for x in boxplot_scores[dataset][scaffolders[i]]]), ')', sep='', file=f)

    print(r'''
col2trans = function(colour) {
    x=as.vector(col2rgb(colour))
    return(rgb(x[1],x[2],x[3],20,maxColorValue=255))
}

''', file=f)

    print('cols=', r_colours_string, file=f)
    print('trans_cols=sapply(cols, col2trans)', file=f)
    #print('svg("', r_script, '.svg")', sep='', file=f)
    print('pdf("', r_script, '.pdf", useDingbats=F)', sep='', file=f)
    print('par(mar=c(9,4,2,2) + 0.1)', file=f)
    print('boxplot(', ','.join(scaffnames), ',',
          #'col=trans_cols,',
          'col=cols,',
          'border="darkslategray",',
          'outline=F,',
          'ylim=c(0,1),',
          'names=', r_scaffnames_string, ',',
          'horizontal=F,',
          'las=2,',
          'ylab="Normalised score"',
          ')', sep='', file=f)
    print('points(', r_stars, ', pch=21, col="black", bg="white", cex=1.5)', sep='', file=f)
    print('dev.off()', file=f)
    utils.close(f)
    utils.syscall('R CMD BATCH ' + r_script)

# print histogram data for R


#for dataset in scores:
#    if 'Staph' not in dataset:
#        continue
#    r_script = options.outprefix + '.score_iter.' + dataset + '.R'
#    f = utils.open_file_write(r_script)
#
#    print('pdf("', r_script, '.pdf")', sep='', file=f)
#    print('plot(xlim=c(0,', bins, '), ylim=c(0,1000), NULL)', file=f)
#
#    for scaff in scaffolders:
#        print('lines(c(',
#              ','.join([str(x) for x in scores[dataset][scaff]]), '),'
#              'col="', r_colours[scaff], '")',
#              sep='', file=f)
#
#    print('dev.off()', file=f)
#
#    utils.close(f)







# make a tsv file of all the stats
#f = utils.open_file_write(options.outprefix + '.stats.tsv')
#print('Scaffolder', 'Good joins',
#      '\t'.join([str(x) for x in possible_flags if x not in [0,16]]),
#      'Bad joins', 'Total joins', '% correct joins', 'Lost tags', 'Skipped tags', 'CPU', 'Mem', 'Extra CPU', 'Extra Mem', sep='\t', file=f)
#
#for scaff in scaffolders:
#    r = results[scaff]
#    fc = r['flag_counts']
#    total_joins = sum([fc[x] for x in possible_flags if x != 16])
#    try:
#        percent_good = round(100 * (1 - float(fc['bad_joins'] / total_joins)), 2)
#    except ZeroDivisionError:
#        percent_good = 0
#
#    print(scaff,
#          '\t'.join([str(fc[x]) for x in possible_flags if x != 16]),
#          fc['bad_joins'],
#          total_joins,
#          percent_good,
#          fc['lost'],
#          fc['skipped'],
#          r['CPU'], r['mem'], sep='\t', file=f)
#
#
#utils.close(f)


# make a scatter plot
#r_script = options.outprefix + '.plot.R'
#f = utils.open_file_write(r_script)
#
#
#x_coords = [results[scaff]['flag_counts'][0] for scaff in scaffolders]
#y_coords = [results[scaff]['flag_counts']['bad_joins'] for scaff in scaffolders]
#x_max = max(x_coords)
#y_max = max(y_coords)
#
#
#for type in ['pdf', 'png']:
#    print(type + '("' + r_script + '.scatter.' + type + '")', file=f)
#
#    print('plot(c(' + ','.join(str(x) for x in x_coords), '), ',
#          'c(', ','.join(str(x) for x in y_coords), '), ',
#          'xlab="Correct joins", ',
#          'ylab="Incorrect joins", ',
#          'xlim=c(0,', x_max, '), ',
#          'ylim=c(0,', y_max, '), ',
#          'col=', r_colour_vector, ', ',
#          'pch=', r_symbol_vector, ', ',
#          'bg=', r_colour_vector,
#          ')', sep='', file=f)
#
#    print(r_legend, file=f)
#
#    print('dev.off()', file=f)
#
#utils.close(f)
#
#
#utils.syscall('R CMD BATCH ' + r_script)


